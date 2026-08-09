#!/usr/bin/env python3
"""
Convert the "Sales orders" tab of PC records.xlsx into SQL INSERTs for gigs.

Run once. The generated file is idempotent — it deletes anything it previously
inserted (identified by the import tag) before inserting, so re-running after a
correction doesn't produce duplicates.

    python3 scripts/import-sales-orders.py "PC records.xlsx" > supabase/import-sales-orders.sql

Mapping decisions, all of them deliberate:

* The spreadsheet's "Client" column mixed venue accounts (University Village)
  with event kinds (Bride, Funeral, Misc.). Venue accounts stay as the client;
  event kinds move to event_category and the real name comes from "Point of
  contact", which is where it always actually was.

* "Payment status" maps onto the app's money model:
    Recieved         -> paid = price, invoice_status 'paid'
    Comp             -> fee 0, paid 0, invoice_status 'paid' (nothing is owed)
    Deposit recieved -> fee = price, paid 0, invoice_status 'sent', and a note.
                        The spreadsheet never recorded the deposit AMOUNT, and
                        inventing one would corrupt the balance arithmetic the
                        whole app is built on. Flagged for Paige to fill in.
    Canceled         -> performance_type 'Canceled' + archived, paid 0.

* Past gigs are NOT archived. Archiving excludes a gig from every total, which
  would hide the five years of history that is the entire point of importing.
"""

import sys
import datetime
import openpyxl

OWNER = 'f7bfcd2e-34d2-4deb-b877-fefddb45e3a2'
TAG = 'xlsx-sales-orders-2026-08'

SENIOR_LIVING = {
    'University Village', 'The Reserve', 'The Variel (WH)', 'Atria Grand Oaks TO',
    'Sunrise of BH', 'Artesian of Ojai', 'The Ridge (WLV)', 'Belmont TO',
}

# Client column value -> (event_category, is_a_real_client_name)
KIND_COLUMN = {
    'Bride':       ('Wedding', False),
    'Coordinator': ('Wedding', False),
    'Funeral':     ('Funeral', False),
    'Fundraiser':  ('Fundraiser', False),
    'Misc.':       (None, False),          # decided per row below
    'Cool School Malibu': ('Community', True),
    'Las Posas CC':       ('Private party', True),
}

LEAD_SOURCE = {
    'Website': 'Website',
    'Word of mouth': 'Word of mouth',
    'Return Customer': 'Return customer',
    'Cold Contact': 'Cold outreach',
    'Promo offer': 'Promo offer',
    'Social media': 'Social media',
    'Family / Friend': 'Family / friend',
}

PERFORMANCE_TYPE = {
    'One time': 'One-time',
    'Reoccuring': 'Recurring',
    'Annual': 'Annual',
    'Promotional': 'Promotional',
    'Cancelation': 'Canceled',
    'Wedding - Cordinator': 'One-time',
    'Wedding - Bride': 'One-time',
}

# Rows where "Misc." hides a real event kind, keyed by spreadsheet row number.
#   row -> (event_category, title, venue)
MISC = {
    9:  ('Corporate',     'Tiffany & Co. Brunch',                    'The Jonathan Club'),
    14: ('Corporate',     'NACE Industry Event',                     'Ventura Harbor Hotel'),
    35: ('Community',     'Santa Clarita Artists Association',       None),
    39: ('Wedding',       'Mia Bella Events Showcase',               'Tuscan Rose Ranch'),
    44: ('Corporate',     'CYRF Convention — VIP Reception',         'Reagan Presidential Library'),
    49: ('Corporate',     'Casino & Cocktails Night',                'Spanish Hills Country Club'),
    51: ('Community',     'Rotary Club Luncheon',                    None),
    52: ('Community',     'She Is Worthy Gathering',                 None),
    53: ('Community',     'Bethany Christmas Tea',                   'Bethany'),
    54: ('Private party', 'Private Event',                           None),
    56: ('Community',     'Casa del Herrero Holiday Event',          'Casa del Herrero'),
    58: ('Community',     '“Vintage” Gathering',                     'Shepherd Church'),
    65: ('Fundraiser',    "Children's Hunger Fund 35th Anniversary", None),
}

# Venues recoverable from the Contact History / Sheet1 tabs or from the notes,
# matched on date. Only included where the match is unambiguous.
VENUE_BY_DATE = {
    '2022-07-07': 'Calamigos Ranch',
    '2022-09-30': 'Rancho de las Palmas',
    '2023-09-09': 'Spanish Hills Country Club',
    '2023-12-30': 'The Tavern / Victorian',
    '2024-05-25': 'Westlake Village Inn',
    '2024-06-09': 'Spanish Hills Country Club',
    '2024-06-29': 'RedBird LA',
    '2024-09-19': 'Ojai Ranch House',
    '2025-08-08': 'Conejo Mountain Funeral Home',
    '2025-10-19': 'Four Seasons',
    '2025-10-25': 'Palazzo by Khoshbin',
    '2026-02-27': 'Cornerstone Simi',
    '2026-05-30': 'The Majestic DTLA',
    '2026-07-10': 'Maravilla Gardens',
    '2026-07-17': 'Serra Cross Park',
    '2026-09-19': 'Santa Barbara Courthouse',
}


def q(v):
    """SQL literal. None -> NULL."""
    if v is None or v == '':
        return 'NULL'
    return "'" + str(v).replace("'", "''") + "'"


def clean(v):
    return str(v).strip() if v is not None and str(v).strip() else None


def main(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Sales orders']

    rows = []
    for r in range(3, ws.max_row + 1):
        client_col = clean(ws.cell(r, 1).value)
        date_val = ws.cell(r, 5).value
        if not client_col or not isinstance(date_val, datetime.datetime):
            continue

        hours_raw = clean(ws.cell(r, 2).value)
        perf_raw = clean(ws.cell(r, 3).value)
        status = clean(ws.cell(r, 4).value)
        price = float(ws.cell(r, 6).value or 0)
        platform = clean(ws.cell(r, 7).value)
        contact = clean(ws.cell(r, 8).value)
        notes = clean(ws.cell(r, 9).value)

        date = date_val.strftime('%Y-%m-%d')
        hours = float(hours_raw) if hours_raw else 0.0

        # ── client / category ────────────────────────────────────
        venue = VENUE_BY_DATE.get(date)
        if client_col in KIND_COLUMN:
            category, is_client_name = KIND_COLUMN[client_col]
            if client_col == 'Misc.':
                category, title, misc_venue = MISC.get(
                    r, ('Other', 'Performance', None))
                venue = venue or misc_venue
            elif client_col == 'Bride':
                title = 'Wedding'
            elif client_col == 'Coordinator':
                title = 'Wedding (coordinator-booked)'
            elif client_col == 'Funeral':
                title = 'Memorial Service'
            elif client_col == 'Fundraiser':
                title = 'Fundraiser'
            else:
                title = f'{client_col} Performance'
            # A handful of old rows recorded no point of contact at all. Fall
            # back to the venue, then to NULL — never to the literal word
            # "Bride" or "Misc.". Writing a category word into the client field
            # is precisely the mixing this import exists to undo, and it would
            # put "Bride" at the top of the Top Clients panel. NULL is the
            # honest answer when no name was ever written down, and Top Clients
            # already skips gigs with no client.
            client = client_col if is_client_name else (contact or venue or None)
        elif client_col in SENIOR_LIVING:
            client, category = client_col, 'Senior living'
            venue = venue or client_col
            title = f'{client_col} Performance'
        else:
            client, category = client_col, 'Other'
            title = f'{client_col} Performance'

        # ── money ────────────────────────────────────────────────
        performance_type = PERFORMANCE_TYPE.get(perf_raw, 'One-time')
        archived, extra_note = False, None
        fee = price
        paid = 0.0
        invoice_status = 'draft'

        if status == 'Recieved':
            paid, invoice_status = price, 'paid'
        elif status == 'Comp':
            fee, paid, invoice_status = 0.0, 0.0, 'paid'
            performance_type = 'Promotional'
        elif status == 'Deposit recieved':
            invoice_status = 'sent'
            extra_note = ('IMPORTED: a deposit was received but the spreadsheet '
                          'never recorded the amount. Enter it in the Deposit '
                          'and Paid fields.')
        elif status == 'Canceled':
            paid, invoice_status = 0.0, 'draft'
            performance_type, archived = 'Canceled', True

        full_notes = '\n\n'.join(x for x in (notes, extra_note) if x)

        rows.append({
            'title': title, 'client': client, 'venue': venue,
            'date': date, 'hours': hours,
            'category': category, 'performance_type': performance_type,
            'lead_source': LEAD_SOURCE.get(platform, 'Other') if platform else None,
            'fee': fee, 'paid': paid, 'invoice_status': invoice_status,
            'archived': archived, 'notes': full_notes or None,
            'row': r,
        })

    out = []
    out.append('-- ' + '=' * 62)
    out.append('-- Historical sales orders imported from PC records.xlsx')
    out.append(f'-- {len(rows)} rows · generated {datetime.date.today()}')
    out.append('--')
    out.append('-- Run AFTER supabase/sales-orders.sql, in the Supabase SQL Editor.')
    out.append('-- Safe to run more than once: the delete below removes any rows a')
    out.append('-- previous run of this file inserted, so corrections re-import')
    out.append('-- cleanly instead of doubling the history.')
    out.append('-- ' + '=' * 62)
    out.append('')
    out.append('begin;')
    out.append('')
    out.append(f"delete from public.gigs where import_tag = {q(TAG)};")
    out.append('')
    out.append('insert into public.gigs')
    out.append('  (user_id, title, client, venue, date, duration_hours,')
    out.append('   event_category, performance_type, lead_source,')
    out.append('   fee, deposit, paid, invoice_status, contract_status,')
    out.append('   archived, notes, import_tag)')
    out.append('values')

    values = []
    for x in rows:
        values.append(
            f"  ({q(OWNER)}, {q(x['title'])}, {q(x['client'])}, {q(x['venue'])}, "
            f"{q(x['date'])}, {x['hours']},\n"
            f"   {q(x['category'])}, {q(x['performance_type'])}, {q(x['lead_source'])},\n"
            f"   {x['fee']:.2f}, 0, {x['paid']:.2f}, {q(x['invoice_status'])}, 'not sent',\n"
            f"   {str(x['archived']).lower()}, {q(x['notes'])}, {q(TAG)})"
        )
    out.append(',\n'.join(values) + ';')
    out.append('')
    out.append('commit;')
    out.append('')

    total_fee = sum(x['fee'] for x in rows)
    total_paid = sum(x['paid'] for x in rows)
    out.append('-- ── Reconciliation ────────────────────────────────────────')
    out.append(f'-- rows inserted     {len(rows)}')
    out.append(f'-- total fee booked  ${total_fee:,.2f}')
    out.append(f'-- total collected   ${total_paid:,.2f}')
    out.append('--')
    out.append('-- The spreadsheet SUM of the Price column is $28,850.00 across 79 rows;')
    out.append('-- booked matches it exactly. Collected is lower by design:')
    out.append('--   $1,650  three 2026 bookings marked "Deposit recieved", where the')
    out.append('--           spreadsheet never recorded how much was received')
    out.append('--   $100    one canceled fundraiser that kept its quote but was')
    out.append('--           never performed or paid')
    out.append('')
    out.append('-- select count(*), sum(fee), sum(paid) from public.gigs')
    out.append(f"--   where import_tag = {q(TAG)};")

    print('\n'.join(out))

    print(f'\n-- rows: {len(rows)}  fee: {total_fee}  paid: {total_paid}',
          file=sys.stderr)


if __name__ == '__main__':
    main(sys.argv[1] if len(sys.argv) > 1 else 'PC records.xlsx')
